
import numpy as np
import pandas as pd
from functools import partial
import os
from qdata.io import IO
import pickle

# env = Environment()
# dc = DataCollector(env)
# data = dc.get_data(['close_','mktcap','gpa','ret12m1m','bm','sue', 'accrual'])
# data = dc.get_data(['close_','mktcap','sue', 'accrual','netopasset','ns','y_next'])
# data = dc.get_data(['mktcap','bm','assetgrowth','roa','roe','gpa','accrual','logme','invtocap','invgrowth','invtoasset','netopasset','ns'])
# data = dc.get_data(['mktcap','ret12m', 'vol250d','y_next'])
# data['vol_adj_ret12m'] = data['ret12m'] / data['vol250d']
# bm = PreProcessing(data[['eval_d','infocode','bm']]).data
# gpa= PreProcessing(data[['eval_d','infocode','gpa']]).data
# mom= PreProcessing(data[['eval_d','infocode','ret12m']]).data
# mom= PreProcessing(data[['eval_d','infocode','ret12m1m']]).data
# size= PreProcessing(data[['eval_d','infocode','mktcap']]).data
# sue= PreProcessing(data[['eval_d','infocode','sue']]).data
# y_next = PreProcessing(data[['eval_d','infocode','y_next']]).data

# model = Strategy(bm, data_for_wgt=size)
# model = Strategy(mom, data_for_wgt=size)
# model = Strategy(size, data_for_wgt=size)
# model.truncate()
# model.make_order_table(multiple=-1.0)
# saver = Saver(model.data)
# saver.export_data_to_db(['wgt'], table_nm='wgt_table', db_name='qpipe')

# model.truncate()
# model.make_order_table_with_double_sort(size)
# model.make_order_table_with_double_sort(size, q=0.33, second_q=0.5, multiple=1.0)
# model.make_order_table_with_double_sort(bm, q=0.5, second_q=0.33, multiple=-1.0)
# saver = Saver(model.data)
# saver.export_data_to_db(['value_','value_for_wgt','adj_data','selection','second_selection','selected_value_for_wgt','wgt'], table_nm='wgt_table', db_name='qpipe')

# dc.get_data_to_db(['mktcap','gpa','roa'], table_nm='my_first_test')








# env = Environment()
# dc = DataCollector(env)
# data = dc.get_data(['mktcap','bm','assetgrowth','roa','roe','gpa','accrual','logme','invtocap','invgrowth','invtoasset','netopasset','ns'])

# size= PreProcessing(data[['eval_d','infocode','mktcap']]).data
# model = Strategy(size, data_for_wgt=size)
# model.truncate(p_upper=1)
# model.join_data(PreProcessing(data[['eval_d','infocode','bm']]).data, 'bm', inplace=True)
# model.join_data(PreProcessing(data[['eval_d','infocode','assetgrowth']]).data, 'assetgrowth', inplace=True)
# model.join_data(PreProcessing(data[['eval_d','infocode','roa']]).data, 'roa', inplace=True)
# model.join_data(PreProcessing(data[['eval_d','infocode','roe']]).data, 'roe', inplace=True)
# model.join_data(PreProcessing(data[['eval_d','infocode','gpa']]).data, 'gpa', inplace=True)
# model.join_data(PreProcessing(data[['eval_d','infocode','accrual']]).data, 'accrual', inplace=True)
# model.join_data(PreProcessing(data[['eval_d','infocode','logme']]).data, 'logme', inplace=True)
# model.join_data(PreProcessing(data[['eval_d','infocode','invtocap']]).data, 'invtocap', inplace=True)
# model.join_data(PreProcessing(data[['eval_d','infocode','invgrowth']]).data, 'invgrowth', inplace=True)
# model.join_data(PreProcessing(data[['eval_d','infocode','invtoasset']]).data, 'invtoasset', inplace=True)
# model.join_data(PreProcessing(data[['eval_d','infocode','netopasset']]).data, 'netopasset', inplace=True)
# model.join_data(PreProcessing(data[['eval_d','infocode','ns']]).data, 'ns', inplace=True)
# df = model.data.copy()
# df = df[np.sum(df.ix[:, 4:].isna(), axis=1) <= 3]
# colnames = ['bm','assetgrowth','roa','roe','gpa','accrual','logme','invtocap','invgrowth','invtoasset','netopasset','ns']
# df.ix[:, colnames] = df.groupby(['eval_d']).apply(lambda x: (x - np.mean(x[~np.isnan(x)])) / np.std(x[~np.isnan(x)])).ix[:, colnames]
# saver = Saver(df)
# saver.export_data_to_pickle('financialdata.pkl')
# f = open('financialdata.pkl', 'rb')
# df = pickle.load(f)


class Environment:
    def __init__(self, univ='equity_us'):
        self.univ = univ
        self.sch = Schedule('1976-01-01', '2018-08-31', type_='end', freq_='m')

    def set_universe(self, univ):
        self.univ = univ

    def set_schedule(self, sch):
        self.sch = sch


class Schedule(IO):
    """
        가져오려는 데이터의 기간/주기 설정 클래스
        필수파라미터
        begin_d: 시작날짜 ('yyyy-mm-dd' 형식)
        end_d: 끝날짜 ('yyyy-mm-dd' 형식)
        type_: 어떤 형식으로 받아올지 결정 (end | start | spec | interval)
        if type_ in (end | start) : freq_ ('y','q','m','w','d')
                    (spec)        : days ([2, 3, ...]   months if needed)
                    (interval)    : freq_, interval (integer)
    """
    def __init__(self, begin_d, end_d, type_=None, **kwargs):
        super().__init__()
        self.code = None
        self.desc_ = None
        self.begin_d = begin_d
        self.end_d = end_d
        self.type_ = type_
        self.set_schedule(**kwargs)

    @staticmethod
    def obj_to_dbformat(obj, default='Null', sep=",", end="'"):
        if isinstance(obj, list):
            return_str = end + sep.join([str(i) for i in obj]) + end
        elif obj is None:
            return_str = default
        else:
            return_str = end + str(obj) + end

        return return_str

    def _gen_code(self, freq_=None, interval=None, days=None, months=None):
        self.sqlm.set_db_name('qinv')
        sql_ = "select qinv.dbo.FS_ScheduleCodeMaker({}, {}, {}, {}, {}, {}, {})".format(
            *tuple(map(Schedule.obj_to_dbformat, [self.begin_d, self.end_d, self.type_, freq_, interval, days, months])))

        code = self.sqlm.db_read(sql_)
        code = code.values[0][0]

        return code

    def set_schedule(self, **kwargs):
        """
        :param kwargs:
            가져오려는 데이터의 기간/주기 설정 클래스
            필수파라미터
            begin_d: 시작날짜 ('yyyy-mm-dd' 형식)
            end_d: 끝날짜 ('yyyy-mm-dd' 형식)
            type_: 어떤 형식으로 받아올지 결정 (end | start | spec | interval)
            if type_ in (end | start) : freq_ ('y','q','m','w','d')
                    (spec)        : days ([2, 3, ...]   months if needed)
                        (interval)    : freq_, interval (integer)
        """
        key_list = [i.lower() for i in kwargs.keys()]
        if 'begin_d' in key_list:
            self.begin_d = kwargs['begin_d']
        if 'end_d' in key_list:
            self.end_d = kwargs['end_d']
        if 'type_' in key_list:
            self.type_ = kwargs['type_']

        if self.type_ in ('end', 'start'):
            assert 'freq_' in key_list
            self.code = self._gen_code(freq_=kwargs['freq_'])
            self.desc_ = "{{'freq_': '{}'}}".format(kwargs['freq_'])
        elif self.type_ == 'spec':
            assert 'days' in key_list
            if 'months' in key_list:
                self.code = self._gen_code(days=kwargs['days'], months=kwargs['months'])
                self.desc_ = "{{'days': {}, 'months': {} }}".format(kwargs['days'], kwargs['months'])
            else:
                self.code = self._gen_code(days=kwargs['days'])
                self.desc_ = "{{'days': {}}}".format(kwargs['days'])
        elif self.type_ == 'interval':
            assert 'freq_' in key_list
            assert 'interval' in key_list
            self.code = self._gen_code(freq_=kwargs['freq_'], interval=kwargs['interval'])
            self.desc_ = "{{'freq_': '{}', 'interval': {}}}".format(kwargs['freq_'], kwargs['interval'])
        else:
            pass


class Asset(IO):
    def __init__(self, asset_type):
        super(__class__, self).__init__()
        self.data_src = dict()
        self.initialize(asset_type)

    def initialize(self, asset_type):
        info_table = self.sqlm.db_read("""select src_nm, db_nm, item_nm_src, data_code_format 
                    from qinv..AssetDataSourceInfo where type_ = '{}' 
                    and src_nm in ('fc','mkt')""".format(asset_type))

        for src in info_table.src_nm:
            db_nm, item_nm_src,  data_code_format = info_table.ix[info_table['src_nm'] == src, 1:].iloc[0]
            self.sqlm.set_db_name(db_nm)
            item_list = [i.lower() for i in list(self.sqlm.db_read(item_nm_src).ix[:, 0])]

            self.data_src[src] = {'item_list': item_list, 'code_format': data_code_format}

    def gen_code(self, item_list):
        return_code = ""
        for src in self.data_src.keys():
            items_in_src = list()
            for item in item_list:
                if item in self.data_src[src]['item_list']:
                    items_in_src.append(item)

            if len(items_in_src) > 0:
                column_str = ", ".join(items_in_src)
                factor_str1 = "'" + "','".join(items_in_src) + "'"
                factor_str2 = "[" + "],[".join(items_in_src) + "]"
                sql_ = self.data_src[src]['code_format']
                sql_ = sql_.replace('[[item_str1]]', factor_str1)
                sql_ = sql_.replace('[[item_str2]]', factor_str2)
                sql_ = sql_.replace('[[column_str]]', column_str)
                sql_ = sql_.replace('[[cond]]', "where eval_d = d.eval_d and infocode = u.infocode")

                return_code += "cross apply ({} ) {}\n".format(sql_, src)

            for item in items_in_src:
                item_list.remove(item)

        if len(item_list) > 0:
            print("Check the following items: {}".format(','.join(item_list)))

        return return_code


class CodeGenerator:
    def __init__(self, univ_code, sch_code):
        self.univ_code = univ_code
        self.sch_code = sch_code

    @staticmethod
    def define_table_structure(item_list):
        table_structure = dict()
        table_structure['columns'] = list()
        table_structure['pk'] = ['eval_d', 'infocode']

        column_list = ['eval_d', 'infocode'] + item_list
        column_type = ['date', 'int'] + ['float'] * len(item_list)
        for column_tuple in zip(column_list, column_type):
            table_structure['columns'].append(column_tuple)

        return table_structure

    @staticmethod
    def get_code_universe(univ):
        return "select type_, code_ from qinv..UniverseList where univ_id='{}'".format(univ)

    @staticmethod
    def gen_skeleton_code():
        sql_ = """
        select *
            from ([[sch_code]]) D
            cross apply ([[univ_code]] and startdate <= d.eval_d and enddate >= d.eval_d) U
            [[item_code]]
            order by d.eval_d, u.infocode
        """
        return sql_

    def gen_code(self, item_list, asset_obj):
        sql_ = CodeGenerator.gen_skeleton_code()
        sql_ = sql_.replace('[[sch_code]]', self.sch_code)
        sql_ = sql_.replace('[[univ_code]]', self.univ_code)
        sql_ = sql_.replace('[[item_code]]', asset_obj.gen_code(item_list))

        return sql_


class Saver(IO):
    def __init__(self, df):
        super(__class__, self).__init__()
        self.df = df

    def export_data_to_db(self, item_list, table_nm='mytest', db_name='qpipe'):
        is_created = self.create_table(table_nm, CodeGenerator.define_table_structure(item_list), db_name=db_name)
        if is_created:
            self.sqlm.set_db_name(db_name)
            self.sqlm.db_insert(self.df.ix[:, ['eval_d', 'infocode'] + item_list], table_name=table_nm, fast_executemany=True)
            return True
        else:
            print("Make sure table name '{}' is correct or drop the table first. "
                  "[USE drop_table(table_nm) method.]".format(table_nm))
            return False

    def export_data_to_file(self, file_nm, dir='./'):
        extension = os.path.splitext(file_nm)[1][1:].lower()
        file_dir = os.path.join(dir, file_nm)
        if extension in ('txt','csv'):
            self.df.to_csv(file_dir, header=True, index=None, sep=',', mode='w')
            # df.to_csv(r'txt//feasible_set.csv', header=True, index=None, sep=',', mode='w')
        elif extension in ('xls', 'xlsx'):
            Saver.df_to_excel(self.df, workbook_nm=file_nm, dir=dir)

    def export_data_to_pickle(self, file_nm, dir='./'):
        file_dir = os.path.join(dir, file_nm)
        f = open(file_dir, 'wb')
        pickle.dump(self.df, f)

    @staticmethod
    def df_to_excel(arr, workbook_nm='untitled.xlsx', worksheet_nm='sheet1', dir='./', is_df=True):
        from openpyxl import Workbook, load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        dir_nm = os.path.join(dir, workbook_nm)
        if not os.path.exists(dir_nm):
            wb = Workbook(dir_nm)
        else:
            wb = load_workbook(dir_nm)

        if worksheet_nm in wb.get_sheet_names():
            std = wb.get_sheet_by_name(worksheet_nm)
            wb.remove_sheet(std)

        ws = wb.create_sheet(worksheet_nm)
        if is_df is True:
            for row in dataframe_to_rows(arr, index=False, header=True):
                ws.append(row)
        else:
            for row in arr:
                ws.append(row.tolist())

        wb.save(filename=dir_nm)


class DataCollector(IO):
    def __init__(self, env):
        super(__class__, self).__init__()
        self.env = env
        self.code_generator = None
        self.asset_obj = None
        self._activate_env()

    def _activate_env(self):
        univ_data = self.sqlm.db_read(CodeGenerator.get_code_universe(self.env.univ))
        assert len(univ_data) == 1, "Check universe name in UniverseList"
        self.asset_obj = Asset(univ_data.type_[0])
        self.code_generator = CodeGenerator(univ_code=univ_data.code_[0], sch_code=self.env.sch.code)

    def set_env(self, env):
        self.env = env
        self._activate_env()

    def get_data(self, item_list):
        sql_ = self.code_generator.gen_code(item_list, self.asset_obj)
        print(sql_)
        return self.sqlm.db_read(sql_)

    def export_data_to_db(self, item_list, table_nm='mytest', db_name='qpipe'):
        is_created = self.create_table(table_nm, CodeGenerator.define_table_structure(item_list), db_name=db_name)
        sql_ = "insert into {}..{} ".format(db_name, table_nm) \
               + self.code_generator.gen_code(item_list, self.asset_obj)
        if is_created:
            self.sqlm.db_execute(sql_)
            return True
        else:
            print("Make sure table name '{}' is correct or drop the table first. "
                  "[USE drop_table(table_nm) method.]".format(table_nm))
            return False

    def export_data_to_file(self, item_list):
        raise NotImplementedError

    def load(self, table_nm, db_name='qpipe'):
        return self.sqlm.db_read("select * from {}..{}".format(db_name, table_nm))


class Calculator:
    @staticmethod
    def winsor_func(arr, p_upper=1, p_lower=0, is_pct=True):
        assert p_upper >= p_lower, "Invalid Boundary."

        if is_pct is True:
            upper_cp = np.percentile(arr, 100. * p_upper)
            lower_cp = np.percentile(arr, 100. * p_lower)
        else:
            upper_cp = p_upper
            lower_cp = p_lower

        ret = arr.copy()
        ret[ret > upper_cp] = upper_cp
        ret[ret < lower_cp] = lower_cp

        return ret

    @staticmethod
    def truncate_func(arr, p_upper=1, p_lower=0, is_pct=True):
        assert p_upper >= p_lower, "Invalid Boundary."

        if is_pct is True:
            upper_cp = np.percentile(arr, 100. * p_upper)
            lower_cp = np.percentile(arr, 100. * p_lower)
        else:
            upper_cp = p_upper
            lower_cp = p_lower

        ret = arr.copy()
        ret[(ret > upper_cp) | (ret < lower_cp)] = np.nan
        # ret[ret < lower_cp] = None

        return ret
        # return arr[(arr <= upper_cp) & (arr >= lower_cp)]


class PreProcessing:
    def __init__(self, data):
        """
        :param data: (eval_d, infocode, value_) format
        """
        assert data.shape[1] == 3, 'Data Format: [eval_d, infocode, value_]'
        data.columns = pd.Index(['eval_d', 'infocode', 'value_'])
        self.sch = data.eval_d.unique()
        self.data = data
        self._delete_na()
        self.truncated = False

    def _delete_na(self):
        self.data = self.data[~self.data.value_.isna()]

    def apply_cross_section(self, ftn, old_column='value_', new_column='value_'):
        self.data[new_column] = self.data.groupby('eval_d')[old_column].transform(ftn)

    def standardize(self):
        self.apply_cross_section(lambda x: (x - np.mean(x)) / np.std(x))

    def truncate(self, p_upper=0.99, p_lower=0.01, is_pct=True):
        if self.truncated:
            print("Already truncated.")
            return False
        truncate_ftn = partial(Calculator.truncate_func, p_upper=p_upper, p_lower=p_lower, is_pct=is_pct)
        self.apply_cross_section(truncate_ftn)
        self._delete_na()
        self.truncated = True
        return True

    def winsorize(self, p_upper=0.99, p_lower=0.01, is_pct=True):
        winsor_ftn = partial(Calculator.winsor_func, p_upper=p_upper, p_lower=p_lower, is_pct=is_pct)
        self.apply_cross_section(winsor_ftn)

    def join_data(self, added_data, column_nm, inplace=False):
        assert added_data.shape[1] == 3, 'Data Format: [eval_d, infocode, value_]'
        added_data.columns = pd.Index(['eval_d', 'infocode', column_nm])
        if inplace is True:
            self.data = pd.merge(self.data, added_data, how='left', on=['eval_d', 'infocode'])
            return self.data
        else:
            return pd.merge(self.data, added_data, how='left', on=['eval_d', 'infocode'])


class Strategy(PreProcessing):
    def __init__(self, data, data_for_wgt=None):
        super(__class__, self).__init__(data)
        self.prepare_for_wgt(data_for_wgt)

    def prepare_for_wgt(self, data_for_wgt=None):
        if data_for_wgt is None:
            self.data['value_for_wgt'] = 1
        else:
            self.data = self.join_data(data_for_wgt, 'value_for_wgt')

    def make_order_table(self, q=0.3, multiple=1.0, long=True, short=True):
        stock_selection_ftn = partial(Strategy.stock_selection_func, q=q, long=long, short=short)
        self.apply_cross_section(stock_selection_ftn, new_column='selection')
        value_to_wgt_ftn = partial(Strategy.value_to_wgt_func, multiple=multiple)
        self.data['selected_value_for_wgt'] = self.data['value_for_wgt'] * self.data['selection']
        self.apply_cross_section(value_to_wgt_ftn, old_column='selected_value_for_wgt', new_column='wgt')
        # self.data.drop(['selection', 'selected_value_for_wgt'])

    def make_order_table_with_double_sort(self, second_data_for_adj, q=0.3, second_q=0.5, multiple=1.0, long=True, short=True):
        self.data = self.join_data(second_data_for_adj, 'adj_data')
        self.data = self.data[~self.data.adj_data.isna()]

        stock_selection_ftn = partial(Strategy.stock_selection_func, q=q, long=long, short=short)
        self.apply_cross_section(stock_selection_ftn, new_column='selection')

        adj_selection_ftn = partial(Strategy.stock_selection_func, q=second_q, long=True, short=True)
        self.apply_cross_section(adj_selection_ftn, old_column='adj_data', new_column='second_selection')

        value_to_wgt_ftn = partial(Strategy.value_to_wgt_func, multiple=multiple * second_q)
        self.data['selected_value_for_wgt'] = self.data['value_for_wgt'] * self.data['selection']
        self.data['wgt'] = self.data.groupby(['eval_d', 'second_selection'])['selected_value_for_wgt'].transform(value_to_wgt_ftn)

    def calc_portfolio_returns(self, y_next):
        data = self.join_data(y_next, 'y_next')
        data['y_wgt'] = data['y_next'] * data['wgt']
        ret = pd.DataFrame({'y': data.groupby(['eval_d'])['y_wgt'].sum()})
        ret['count'] = data[(data['wgt'] > 0) | (data['wgt'] < 0)].groupby(['eval_d'])['infocode'].count()
        ret = ret[ret['count'] >= 50]
        ret.index = pd.to_datetime(ret.index)
        return ret

    @staticmethod
    def value_to_wgt_func(arr, multiple=1.0):
        ret = arr.copy()
        ret[ret > 0] = ret[ret > 0] / np.sum(ret[ret > 0])
        ret[ret < 0] = ret[ret < 0] / np.abs(np.sum(ret[ret < 0]))
        return ret * multiple

    @staticmethod
    def stock_selection_func(arr, q, long=True, short=True):
        # q가 0.5보다 크고 long/short 모두 True면, 겹치는 부분은 값이 상쇄되어 1-q와 동일효과
        if long is True:
            lower_long = np.percentile(arr, 100. * (1 - q))

        if short is True:
            upper_short = np.percentile(arr, 100. * q)

        ret = np.zeros_like(arr)
        ret[(arr >= lower_long)] += 1
        ret[(arr < upper_short)] -= 1
        return ret


class Tester:
    def __init__(self, data):
        self.data = data

    def summary(self):
        begin_d = min(self.data['eval_d'])
        end_d = max(self.data['eval_d'])
        total_firm = len(self.data['infocode'].unique())






# env = Environment()
# dc = DataCollector(env)
# rs = dc.sqlm.db_read('select * from qinv..equityreturnseriesmonthly')
# rs_notnull = rs[~rs.y_next.isna()]
# a =rs_notnull.groupby(['eval_d'])[['0','1','2','3','4','5','6','7','8','9','10','11','12','13','14']].apply(lambda x: (x - np.mean(x[~np.isnan(x)]))/ np.std(x[~np.isnan(x)]))
# b =rs_notnull.groupby(['eval_d'])['y_next'].apply(lambda x: (x >= np.percentile(x[~np.isnan(x)], 80)) + (x <= np.percentile(x[~np.isnan(x)], 20)) * (-1))
#
# # df.ix[:, colnames] = df.groupby(['eval_d']).apply(lambda x: (x - np.mean(x[~np.isnan(x)])) / np.std(x[~np.isnan(x)])).ix[:, colnames]
#
# data = np.array(a)
# data[np.isnan(data)] = 0
# label = np.concatenate([np.array((b == 1) * 1).reshape(-1, 1),
#                         np.array((b == 0) * 1).reshape(-1, 1),
#                         np.array((b == -1) * 1).reshape(-1, 1)], axis=1)
#
#
# np.random.seed(1234)
# shuffled = np.random.choice(data.shape[0], data.shape[0], replace=False)
#
<<<<<<< HEAD
# train_data = data[shuffled][:int(len(shuffled) * 0.6), :]
# train_label = label[shuffled][:int(len(shuffled) * 0.6), :]
#
# valid_data = data[shuffled][int(len(shuffled) * 0.6):int(len(shuffled) * 0.8), :]
# valid_label = label[shuffled][int(len(shuffled) * 0.6):int(len(shuffled) * 0.8), :]
=======
# train_data = data[shuffled][:int(len(shuffled) * 0.8), :]
# train_label = label[shuffled][:int(len(shuffled) * 0.8), :]
>>>>>>> origin
#
# test_data = data[shuffled][int(len(shuffled) * 0.8):, :]
# test_label = label[shuffled][int(len(shuffled) * 0.8):, :]
#
#
# example_data = data[shuffled][:1000, :]
# example_label = label[shuffled][:1000, :]
#
#
# import tensorflow as tf
# from keras.layers import Input, Dense
# from keras.models import Model
<<<<<<< HEAD
# import keras
=======
>>>>>>> origin
#
# # This returns a tensor
# inputs = Input(shape=(15,))
#
# # a layer instance is callable on a tensor, and returns a tensor
<<<<<<< HEAD
#
# x = Dense(30,
#           kernel_initializer='glorot_normal',
#           activation='relu')(inputs)
# x = Dense(30,
#           kernel_initializer='glorot_normal',
#           activation='relu')(x)
=======
# x = Dense(30, activation='relu')(inputs)
# x = Dense(20, activation='relu')(x)
>>>>>>> origin
# predictions = Dense(3, activation='softmax')(x)
#
# def my_accuracy(y_true, y_pred):
#     pred_ = tf.argmax(y_pred, 1)
#     equality = tf.equal(pred_, tf.argmax(y_true, 1))
#     return tf.reduce_mean(tf.cast(equality, tf.float32))
#
# # This creates a model that includes
# # the Input layer and three Dense layers
# model = Model(inputs=inputs, outputs=predictions)
# model.compile(optimizer='adam',
#               loss='categorical_crossentropy',
<<<<<<< HEAD
#               metrics=['acc'])
#
#
# tb_acc = keras.callbacks.TensorBoard(log_dir="./graph", histogram_freq=0, write_graph=True, write_images=True)
# model.fit(train_data, train_label, epochs=1000, batch_size=10000, validation_data=(valid_data, valid_label),
#           callbacks=[tb_acc])  # starts training
# score = model.evaluate(example_data, example_label, batch_size=128)
# model.predict(example_data)
#
#
#
#
# import keras
# from keras.models import Sequential
# from keras.layers import LSTM, BatchNormalization, Dense
# import keras.backend as K
#
# data_dim = 1
# timesteps = 15
# nb_classes = 3
#
# model2 = Sequential()
# model2.add(LSTM(50,
#                 return_sequences=True,
#                 input_shape=(timesteps, data_dim)))
# model2.add(BatchNormalization())
# model2.add(LSTM(50))
# model2.add(BatchNormalization())
# model2.add(Dense(3, activation='softmax'))
# model2.compile(optimizer=keras.optimizers.Adam(lr=0.0001),
#               loss='categorical_crossentropy',
#               metrics=['acc'])
#
#
# example_data = data[shuffled][:10000, :]
# example_data_lstm = example_data.reshape((-1, timesteps, data_dim))
# example_label = label[shuffled][:10000, :]
#
# tb_acc = keras.callbacks.TensorBoard(log_dir="./graph", histogram_freq=0, write_graph=True, write_images=True)
# model2.fit(example_data_lstm, example_label, epochs=5, batch_size=100,
#           callbacks=[tb_acc])  # starts training
#
# train_data_lstm = train_data.reshape((-1, timesteps, data_dim))
# valid_data_lstm = valid_data.reshape(-1, timesteps, data_dim)
# tb_acc = keras.callbacks.TensorBoard(log_dir="./graph", histogram_freq=0, write_graph=True, write_images=True)
# model2.fit(train_data_lstm, train_label, epochs=100, batch_size=2000, validation_data=(valid_data_lstm, valid_label),
#           callbacks=[tb_acc])  # starts training



















###########
=======
#               metrics=['acc', my_accuracy])
# model.fit(example_data, example_label, epochs=20, batch_size=30)  # starts training
# score = model.evaluate(example_data, example_label, batch_size=128)
# model.predict(example_data)
>>>>>>> origin

#
# import pickle
# f = open('financialdata.pkl', 'rb')
# fin_data = pickle.load(f)
#
<<<<<<< HEAD
# from keras.utils import to_categorical
#
=======
#
# from keras.utils import to_categorical
>>>>>>> origin
#
# env = Environment()
# dc = DataCollector(env)
# rs = dc.sqlm.db_read('select * from qinv..equityreturnseriesmonthly')
# rs_notnull = rs[~rs.y_next.isna()]
<<<<<<< HEAD
# # a =pd.merge(rs_notnull[['eval_d', 'infocode']],
=======
# # a = pd.merge(rs_notnull[['eval_d', 'infocode', 'y_next']],
>>>>>>> origin
# #             rs_notnull.groupby(['eval_d'])[['0','1','2','3','4','5','6','7','8','9','10','11','12','13','14']].apply(
# #                 lambda x: (x - np.mean(x[~np.isnan(x)]))/ np.std(x[~np.isnan(x)])),
# #             left_index=True,
# #             right_index=True)
#
# a = rs_notnull.copy()
# a['y_port'] = rs_notnull.groupby(['eval_d'])['y_next'].apply(
#     lambda x: (x >= np.percentile(x[~np.isnan(x)], 10)) * 1.
<<<<<<< HEAD
#               + (x >= np.percentile(x[~np.isnan(x)], 20)) * 1.
#               + (x >= np.percentile(x[~np.isnan(x)], 30)) * 1.
#               + (x >= np.percentile(x[~np.isnan(x)], 40)) * 1.
#               + (x >= np.percentile(x[~np.isnan(x)], 50)) * 1.
#               + (x >= np.percentile(x[~np.isnan(x)], 60)) * 1.
#               + (x >= np.percentile(x[~np.isnan(x)], 70)) * 1.
#               + (x >= np.percentile(x[~np.isnan(x)], 80)) * 1.
#               + (x >= np.percentile(x[~np.isnan(x)], 90)) * 1.)
#
# # a['y_next'] = rs_notnull.groupby(['eval_d'])['y_next'].apply(
# #     lambda x: (x >= np.percentile(x[~np.isnan(x)], 80)) + (x <= np.percentile(x[~np.isnan(x)], 20)) * (-1))
#
# fc_list = ['bm','assetgrowth','roa','roe','gpa','accrual','logme','invtocap','invgrowth','invtoasset','netopasset','ns']
# y_list = ['11','10','9','8','7','6','5','4','3','2','1','0']
# x = pd.merge(fin_data, a, how='inner', on=['eval_d','infocode'])
# data = x.dropna(axis=0, how='any')
# data_f = np.array(data[fc_list])
# # data_p = np.array(data[['14','13','12','11','10','9','8','7','6','5','4','3','2','1','0']])
# # data_f = np.array(data[['bm','assetgrowth','gpa','accrual','logme','invtocap','invgrowth','invtoasset','netopasset','ns']])
# data_p = np.array(data[y_list])
# # data_label = np.array(data[['y_next']])
=======
#                 + (x >= np.percentile(x[~np.isnan(x)], 20)) * 1.
#                 + (x >= np.percentile(x[~np.isnan(x)], 30)) * 1.
#                 + (x >= np.percentile(x[~np.isnan(x)], 40)) * 1.
#                 + (x >= np.percentile(x[~np.isnan(x)], 50)) * 1.
#                 + (x >= np.percentile(x[~np.isnan(x)], 60)) * 1.
#                 + (x >= np.percentile(x[~np.isnan(x)], 70)) * 1.
#                 + (x >= np.percentile(x[~np.isnan(x)], 80)) * 1.
#                 + (x >= np.percentile(x[~np.isnan(x)], 90)) * 1.)
#
#
# # a['y_next'] = rs_notnull.groupby(['eval_d'])['y_next'].apply(
# #     lambda x: (x >= np.percentile(x[~np.isnan(x)], 66)) + (x <= np.percentile(x[~np.isnan(x)], 33)) * (-1))
#
# x = pd.merge(fin_data, a, how='inner', on=['eval_d','infocode'])
# data = x.dropna(axis=0, how='any')
# data_f = np.array(data[['bm','assetgrowth','roa','roe','gpa','accrual','logme','invtocap','invgrowth','invtoasset','netopasset','ns']])
# data_p = np.array(data[['11','10','9','8','7','6','5','4','3','2','1','0']])
# # data_p = np.array(data[['14','13','12','11','10','9','8','7','6','5','4','3','2','1','0']])
# # data_label = np.array(data[['y_next']])
#
>>>>>>> origin
# data_label = to_categorical(data[['y_port']])
#
# np.random.seed(1234)
# shuffled = np.random.choice(data_label.shape[0], data_label.shape[0], replace=False)
#
# train_data_p = data_p[shuffled][:int(len(shuffled) * 0.6), :]
# train_data_f = data_f[shuffled][:int(len(shuffled) * 0.6), :]
# train_label = data_label[shuffled][:int(len(shuffled) * 0.6), :]
#
# valid_data_p = data_p[shuffled][int(len(shuffled) * 0.6):int(len(shuffled) * 0.8), :]
# valid_data_f = data_f[shuffled][int(len(shuffled) * 0.6):int(len(shuffled) * 0.8), :]
# valid_label = data_label[shuffled][int(len(shuffled) * 0.6):int(len(shuffled) * 0.8), :]
#
# test_data_p = data_p[shuffled][int(len(shuffled) * 0.8):, :]
# test_data_f = data_f[shuffled][int(len(shuffled) * 0.8):, :]
# test_label = data_label[shuffled][int(len(shuffled) * 0.8):, :]
#
#
<<<<<<< HEAD
#
# ### test1 start
# import keras
# from keras.models import Sequential, Model
# from keras.layers import LSTM, Dense, Input, BatchNormalization, LeakyReLU, RepeatVector
# from keras.callbacks import EarlyStopping, TensorBoard
# import keras.backend as K
#
# timesteps = 12
# tb_acc = TensorBoard(log_dir="./graph", histogram_freq=0, write_graph=True, write_images=True)
# early_stopping = EarlyStopping(patience=5)
#
# y_bm = pd.DataFrame(columns=['eval_d', 'y'])
# bm = a.groupby('eval_d')['y_next'].mean()
# date_ = a.eval_d.unique()
# a['cum_y'] = np.log(1 + a['1']) + np.log(1 + a['2']) + np.log(1 + a['3']) + np.log(1 + a['4']) \
#              + np.log(1 + a['5']) + np.log(1 + a['6']) + np.log(1 + a['7']) + np.log(1 + a['8']) \
#              + np.log(1 + a['9']) + np.log(1 + a['10']) + np.log(1 + a['11'])
=======
# import keras
# from keras.models import Sequential, Model
# from keras.layers import LSTM, Dense, Input, Flatten, BatchNormalization, LeakyReLU, RepeatVector
# from keras.callbacks import EarlyStopping, TensorBoard
# import keras.backend as K
# from qtest.nalu import NALU
#
#
# data_dim = 1
# lstm_size = 128
# timesteps = 12
# n_fc = 12
# nb_classes = 3
# train_data_p_lstm = train_data_p.reshape((-1, timesteps, data_dim))
# valid_data_p_lstm = valid_data_p.reshape(-1, timesteps, data_dim)
#
# tb_acc = TensorBoard(log_dir="./graph", histogram_freq=0, write_graph=True, write_images=True)
# early_stopping = EarlyStopping(patience=5)
#
#
# input_p = Input(shape=(timesteps, data_dim), name='input_p')
# input_f = Input(shape=(n_fc,), name='input_f')
#
#
#
#
#
# # test start
#
# y_bm = pd.DataFrame(columns=['eval_d', 'y'])
# y_bm30 = pd.DataFrame(columns=['eval_d', 'y'])
# bm = a.groupby('eval_d')['y_next'].mean()
# date_ = a.eval_d.unique()
# a['cum_y'] = np.log(1 + a['1']) + np.log(1 + a['2']) + np.log(1 + a['3']) + np.log(1 + a['4']) + np.log(1 + a['5'])\
#              +np.log(1 + a['6']) + np.log(1 + a['7']) + np.log(1 + a['8']) + np.log(1 + a['9']) + np.log(1 + a['10']) + np.log(1 + a['11'])
#
# for i in range(120, len(date_)):
#     data_test = a[a.eval_d == date_[i]]
#     data_test['y_next_port'] = (data_test.cum_y >= np.percentile(data_test.cum_y, 90)) * 1. / np.sum(data_test.cum_y >= np.percentile(data_test.cum_y, 90)) \
#                                + (data_test.cum_y <= np.percentile(data_test.cum_y, 10)) * -1. / np.sum(data_test.cum_y <= np.percentile(data_test.cum_y, 10))
#
#     data_test['y_next_port30'] = (data_test.cum_y >= np.percentile(data_test.cum_y, 70)) * 1. / np.sum(data_test.cum_y >= np.percentile(data_test.cum_y, 70)) \
#                                + (data_test.cum_y <= np.percentile(data_test.cum_y, 30)) * -1. / np.sum(data_test.cum_y <= np.percentile(data_test.cum_y, 30))
#
#     y_bm = y_bm.append({'eval_d': date_[i], 'y': np.sum(data_test.y_next * data_test.y_next_port)}, ignore_index=True)
#     y_bm30 = y_bm.append({'eval_d': date_[i], 'y': np.sum(data_test.y_next * data_test.y_next_port30)}, ignore_index=True)
#
# y_bm.to_csv("bm.csv")
# y_bm30.to_csv("bm30.csv")
#
>>>>>>> origin
#
#
# def to_ordinal_categorical(df):
#     n_class = df.nunique()
#     ret = np.zeros([len(df), n_class])
#
#     for i in range(n_class):
#         ret[np.array(df == i).reshape([-1])] = np.array([1 if j >= i else 0 for j in range(n_class)])
#
#     return ret
#
<<<<<<< HEAD
# y_table = pd.DataFrame(columns=['eval_d', 'y_long', 'y_short'])
# y_table30 = pd.DataFrame(columns=['eval_d', 'y'])
# for i, t in enumerate(range(240, len(date_))):
#     print(date_[t])
#     data_selected_temp = a[(a.eval_d > date_[t - 240]) & (a.eval_d <= date_[t])]
#     data_selected = data_selected_temp.dropna(axis=0, how='any')
#     if t % 12 == 9 or i == 0:
#         data_train = data_selected[data_selected.eval_d <= date_[t - 48]]
#         data_arr_train = np.array(data_train[['11', '10', '9','8','7','6','5','4','3','2','1','0']])
#         label_arr_train = to_ordinal_categorical(data_train['y_port'])
#
#         data_valid = data_selected[(data_selected.eval_d > date_[t - 48]) & (data_selected.eval_d <= date_[t - 12])]
#         data_arr_valid = np.array(data_valid[['11', '10', '9','8','7','6','5','4','3','2','1','0']])
#         label_arr_valid = to_ordinal_categorical(data_valid['y_port'])
#
#         input_ = Input(shape=(timesteps, ), name='input')
=======
#
#
# y_table = pd.DataFrame(columns=['eval_d', 'y'])
# y_table30 = pd.DataFrame(columns=['eval_d', 'y'])
#
# date_ = a.eval_d.unique()
# for i in range(240, len(date_)):
#     print(date_[i])
#     data_selected_temp = a[(a.eval_d > date_[i - 240]) & (a.eval_d <= date_[i])]
#     data_selected = data_selected_temp.dropna(axis=0, how='any')
#     if i % 12 == 9 or i == 240:
#         data_train = data_selected[data_selected.eval_d <= date_[i-48]]
#         data_arr_train = np.array(data_train[['11', '10', '9', '8', '7', '6', '5', '4', '3', '2', '1', '0']])
#         label_arr_train = to_ordinal_categorical(data_train[['y_port']])
#
#         data_valid = data_selected[(data_selected.eval_d > date_[i-48]) & (data_selected.eval_d <= date_[i-12])]
#         data_arr_valid = np.array(data_valid[['11', '10', '9', '8', '7', '6', '5', '4', '3', '2', '1', '0']])
#         label_arr_valid = to_ordinal_categorical(data_valid[['y_port']])
#
#         input_ = Input(shape=(timesteps,), name='input')
>>>>>>> origin
#
#         x = Dense(64, activation='relu')(input_)
#         x = BatchNormalization()(x)
#         x = Dense(64, activation='relu')(x)
#         x = BatchNormalization()(x)
#         x = Dense(64, activation='relu')(x)
#         x = BatchNormalization()(x)
#         x = Dense(64, activation='relu')(x)
<<<<<<< HEAD
#         output_ = Dense(10, activation='softmax')(x)
#
#         model = Model(inputs=input_, outputs=output_)
#         model.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['acc'])
#         model.fit(data_arr_train, label_arr_train,
#                   epochs=100,
#                   batch_size=1024,
#                   validation_data=[data_arr_valid, label_arr_valid],
#                   shuffle=True,
#                   callbacks=[early_stopping])
#
#     data_test = data_selected[data_selected.eval_d == date_[t]]
#     data_arr_test = np.array(data_test[['11', '10', '9','8','7','6','5','4','3','2','1','0']])
#
#     predicted = model.predict(data_arr_test)
#     prob = np.sum(predicted[:, 7:], axis=1) - np.sum(predicted[:, :3], axis=1)
#     data_test['y_next_port_l'] = (prob >= np.percentile(prob, 90)) * 1. / np.sum(prob >= np.percentile(prob, 90))
#     data_test['y_next_port_s'] = (prob <= np.percentile(prob, 10)) * 1. / np.sum(prob <= np.percentile(prob, 10))
#
#     y_table = y_table.append({'eval_d': date_[t],
#                               'y_long': np.sum(data_test.y_next * data_test.y_next_port_l),
#                               'y_short': np.sum(data_test.y_next * data_test.y_next_port_s)},
#                              ignore_index=True)
#
# y_table.to_csv("result.csv")
#
#
# ### test1 end
#
#
=======
#         x = BatchNormalization()(x)
#         x = Dense(64, activation='relu')(x)
#         output_ = Dense(10, activation='softmax')(x)
#
#         model = Model(inputs=input_, outputs=output_)
#         model.compile(optimizer='adam',
#                       loss='categorical_crossentropy',
#                       metrics=['acc'])
#
#         model.fit(data_arr_train, label_arr_train,
#               epochs=100,
#               batch_size=1024,date_
#               validation_data=[data_arr_valid, label_arr_valid],
#               shuffle=True,
#               callbacks=[early_stopping])
#
#     data_test = data_selected[data_selected.eval_d == date_[i]]
#     data_arr_test = np.array(data_test[['11', '10', '9', '8', '7', '6', '5', '4', '3', '2', '1', '0']])
#
#     predicted = model.predict(data_arr_test)
#     prob = np.sum(predicted[:, 7:], axis=1) - np.sum(predicted[:, :3], axis=1)
#     data_test['y_next_port'] = (prob >= np.percentile(prob, 90)) * 1. / np.sum(prob >= np.percentile(prob, 90)) \
#                                + (prob <= np.percentile(prob, 10)) * -1. / np.sum(prob <= np.percentile(prob, 10))
#
#     y_table = y_table.append({'eval_d': date_[i], 'y': np.sum(data_test.y_next * data_test.y_next_port)}, ignore_index=True)
#
#
#     data_test['y_next_port30'] = (prob >= np.percentile(prob, 70)) * 1. / np.sum(prob >= np.percentile(prob, 70)) \
#                                + (prob <= np.percentile(prob, 30)) * -1. / np.sum(prob <= np.percentile(prob, 30))
#
#     y_table30 = y_table30.append({'eval_d': date_[i], 'y': np.sum(data_test.y_next * data_test.y_next_port30)}, ignore_index=True)
#
#
# y_table.to_csv("result.csv")
# y_table30.to_csv("result30.csv")
>>>>>>> origin
#
#
#
#
#
#
#
#
<<<<<<< HEAD
#
#
#
#
#
# # test2 start
#
# import keras
# from keras.models import Sequential, Model
# from keras.layers import LSTM, Dense, Input, BatchNormalization, Dropout, LeakyReLU
# import keras.backend as K
# from keras.optimizers import SGD
#
# data_dim = 1
# lstm_size = 30
# timesteps = data_p.shape[1]
# n_fc = data_f.shape[1]
# nb_classes = 3
#
# feature_dropout_rate = 0.2
# dropout_rate = 0.5
#
#
# model_p = Sequential()
# model_p.add(BatchNormalization())
# model_p.add(LSTM(lstm_size, return_sequences=True))# , input_shape=(timesteps, data_dim)))
# model_p.add(LSTM(lstm_size))
# # model_p.add(Flatten())
#
# input_p = Input(shape=(timesteps, data_dim), name='input_p')
# output_p = model_p(input_p)
#
# output_aux = Dense(nb_classes, activation='softmax', name='output_aux')(output_p)
#
#
# input_f = Input(shape=(n_fc,), name='input_f')
# merged = keras.layers.concatenate([input_f, output_p])
#
# model_final = Sequential()
# model_final.add(BatchNormalization())
# # model_final.add(Dropout(feature_dropout_rate))
# model_final.add(Dense(128, activation='linear'))
# model_final.add(LeakyReLU(alpha=0.01))
# # model_final.add(Dropout(dropout_rate))
# model_final.add(BatchNormalization())
# model_final.add(Dense(64, activation='linear'))
# model_final.add(LeakyReLU(alpha=0.01))
# # model_final.add(Dropout(dropout_rate))
# model_final.add(BatchNormalization())
# model_final.add(Dense(32, activation='linear'))
# model_final.add(LeakyReLU(alpha=0.01))
# # model_final.add(BatchNormalization())
# # model_final.add(Dense(3, activation='softmax', name='output'))
#
# output = Dense(3, activation='softmax', name='output')(merged)
# # output = model_final(merged)
#
# model = Model(inputs=[input_p, input_f], outputs=[output, output_aux])
#
# sgd = SGD(lr=0.001, decay=1e-6, momentum=0.9, nesterov=True)
# model.compile(optimizer=sgd,
#               loss={'output': 'categorical_crossentropy', 'output_aux': 'categorical_crossentropy'},
#               metrics=['acc'],
#               loss_weights={'output': 1., 'output_aux': 0.2})
#
# train_data_p_lstm = train_data_p.reshape((-1, timesteps, data_dim))
# valid_data_p_lstm = valid_data_p.reshape(-1, timesteps, data_dim)
# tb_acc = keras.callbacks.TensorBoard(log_dir="./graph", histogram_freq=0, write_graph=True, write_images=True)
# model.fit({'input_p': train_data_p_lstm, 'input_f': train_data_f},
#           {'output': train_label, 'output_aux': train_label}, epochs=100, batch_size=10000,
#           validation_data=([valid_data_p_lstm, valid_data_f], [valid_label, valid_label]),
#           callbacks=[tb_acc])  # starts training
#
# # test2 end
#
#
#
#
#
#
# # test3 start
# import keras
# from keras.models import Sequential, Model
# from keras.layers import LSTM, Dense, Input, Flatten, BatchNormalization, LeakyReLU, RepeatVector
# from keras.callbacks import EarlyStopping, TensorBoard
# import keras.backend as K
# from qtest.nalu import NALU
#
#
# data_dim = 1
# lstm_size = 128
# timesteps = 12
# n_fc = 12
# nb_classes = 3
# train_data_p_lstm = train_data_p.reshape((-1, timesteps, data_dim))
# valid_data_p_lstm = valid_data_p.reshape(-1, timesteps, data_dim)
#
# with K.tf.device('/gpu:0'):
#     tb_acc = TensorBoard(log_dir="./graph", histogram_freq=0, write_graph=True, write_images=True)
#     early_stopping = EarlyStopping(patience=5)
#
#
#     input_p = Input(shape=(timesteps, data_dim), name='input_p')
#     input_f = Input(shape=(n_fc,), name='input_f')
#
#
#
#     encoded_lstm = LSTM(50, return_sequences=True)(input_p)
#     encoded_lstm = LSTM(50)(encoded_lstm)
#
#     decoded_lstm = RepeatVector(timesteps)(encoded_lstm)
#     decoded_lstm = LSTM(50, return_sequences=True)(decoded_lstm)
#     decoded_lstm = LSTM(data_dim, return_sequences=True)(decoded_lstm)
#
#     model_enc_lstm = Model(inputs=input_p, outputs=encoded_lstm)
#     model_dec_lstm = Model(inputs=input_p, outputs=decoded_lstm)
#
#     model_dec_lstm.compile(optimizer='rmsprop',
#                   loss='mean_squared_error')
#
#     model_dec_lstm.fit(train_data_p_lstm, train_data_p_lstm,
#               epochs=100,
#               batch_size=1024,
#               validation_data=[valid_data_p_lstm, valid_data_p_lstm],
#               callbacks=[tb_acc, early_stopping])  # starts training
#
#
# with K.tf.device('/gpu:0'):
#     encoded = Dense(64, activation='relu')(input_f)
#     encoded = Dense(32, activation='relu')(encoded)
#     decoded = Dense(64, activation='relu')(encoded)
#     decoded = NALU(n_fc)(decoded)
#
#
#     model_enc = Model(inputs=input_f, outputs=encoded)
#     model_dec = Model(inputs=input_f, outputs=decoded)
#
#     model_dec.compile(optimizer='rmsprop',
#                   loss='mean_squared_error')
#
#     model_dec.fit(train_data_f, train_data_f,
#               epochs=100,
#               batch_size=1024,
#               validation_data=[valid_data_f, valid_data_f],
#               callbacks=[tb_acc, early_stopping])  # starts training
#
#
# with K.tf.device('/gpu:0'):
#     merged = keras.layers.concatenate([model_enc_lstm(input_p), model_enc(input_f)])
#
#     model_final = Sequential()
#     model_final.add(Dense(128, activation='linear', kernel_initializer='glorot_normal'))
#     model_final.add(LeakyReLU())
#     # model_final.add(NALU(128))
#     model_final.add(BatchNormalization())
#     model_final.add(Dense(64, activation='linear', kernel_initializer='glorot_normal'))
#     model_final.add(LeakyReLU())
#     # model_final.add(NALU(64))
#     model_final.add(BatchNormalization())
#     model_final.add(Dense(32, activation='linear', kernel_initializer='glorot_normal'))
#     model_final.add(LeakyReLU())
#     # model_final.add(NALU(32))
#
#     main_output = Dense(10, activation='softmax', name='output')(model_final(merged))
#
#     model = Model(inputs=[input_p, input_f], outputs=main_output)
#
#     model.compile(optimizer='adam',
#                   loss='categorical_crossentropy',
#                   metrics=['acc'])
#
#
#     model.fit([train_data_p_lstm, train_data_f], train_label,
#               epochs=500,
#               batch_size=2048,
#               validation_data=[[valid_data_p_lstm, valid_data_f], valid_label],
#               callbacks=[tb_acc, early_stopping])  # starts training
#
#
#     y_table = pd.DataFrame(columns=['eval_d', 'y_long', 'y_short'])
#     date_ = data.eval_d.unique()
#     for d in date_:
#         data_test = data[data.eval_d == d]
#         if len(data_test) <= 20:
#             continue
#         print(d)
#
#         data_arr_p = np.array(data_test[y_list]).reshape((-1, timesteps, data_dim))
#         data_arr_f = np.array(data_test[fc_list])
#         predicted = model.predict([data_arr_p, data_arr_f])
#         prob = np.sum(predicted[:, 7:], axis=1) - np.sum(predicted[:, :3], axis=1)
#
#         data_test['y_next_port_l'] = (prob >= np.percentile(prob, 90)) * 1. / np.sum(prob >= np.percentile(prob, 90))
#         data_test['y_next_port_s'] = (prob <= np.percentile(prob, 10)) * 1. / np.sum(prob <= np.percentile(prob, 10))
#
#         y_table = y_table.append({'eval_d': d,
#                                   'y_long': np.sum(data_test.y_next * data_test.y_next_port_l),
#                                   'y_short': np.sum(data_test.y_next * data_test.y_next_port_s)},
#                                  ignore_index=True)
#
#     y_table.to_csv("test_insample.csv")
#
#
# # test3 end
=======
# encoded_lstm = LSTM(50, return_sequences=True)(input_p)
# encoded_lstm = LSTM(50)(encoded_lstm)
#
# decoded_lstm = RepeatVector(timesteps)(encoded_lstm)
# decoded_lstm = LSTM(50, return_sequences=True)(decoded_lstm)
# decoded_lstm = LSTM(data_dim, return_sequences=True)(decoded_lstm)
#
# model_enc_lstm = Model(inputs=input_p, outputs=encoded_lstm)
# model_dec_lstm = Model(inputs=input_p, outputs=decoded_lstm)
#
# model_dec_lstm.compile(optimizer='rmsprop',
#               loss='mean_squared_error')
#
#
#
# model_dec_lstm.fit(train_data_p_lstm, train_data_p_lstm,
#           epochs=100,
#           batch_size=1024,
#           validation_data=[valid_data_p_lstm, valid_data_p_lstm],
#           callbacks=[tb_acc, early_stopping])  # starts training
#
#
#
#
# encoded = Dense(64, activation='relu')(input_f)
# encoded = Dense(32, activation='relu')(encoded)
# decoded = Dense(64, activation='relu')(encoded)
# decoded = NALU(n_fc)(decoded)
#
#
# model_enc = Model(inputs=input_f, outputs=encoded)
# model_dec = Model(inputs=input_f, outputs=decoded)
#
# model_dec.compile(optimizer='rmsprop',
#               loss='mean_squared_error')
#
# model_dec.fit(train_data_f, train_data_f,
#           epochs=100,
#           batch_size=1024,
#           validation_data=[valid_data_f, valid_data_f],
#           callbacks=[early_stopping])  # starts training
#
#
# merged = keras.layers.concatenate([model_enc_lstm(input_p), model_enc(input_f)])
#
# model_final = Sequential()
# model_final.add(Dense(128, activation='linear', kernel_initializer='glorot_normal'))
# model_final.add(LeakyReLU())
# # model_final.add(NALU(128))
# model_final.add(BatchNormalization())
# model_final.add(Dense(64, activation='linear', kernel_initializer='glorot_normal'))
# model_final.add(LeakyReLU())
# # model_final.add(NALU(64))
# model_final.add(BatchNormalization())
# model_final.add(Dense(32, activation='linear', kernel_initializer='glorot_normal'))
# model_final.add(LeakyReLU())
# # model_final.add(NALU(32))
#
# main_output = Dense(3, activation='softmax', name='output')(model_final(merged))
#
# model = Model(inputs=[input_p, input_f], outputs=main_output)
#
# model.compile(optimizer='adam',
#               loss='categorical_crossentropy',
#               metrics=['acc'])
#
#
# model.fit([train_data_p_lstm, train_data_f], train_label,
#           epochs=500,
#           batch_size=2048,
#           validation_data=[[valid_data_p_lstm, valid_data_f], valid_label],
#           callbacks=[tb_acc, early_stopping])  # starts training
#
#
#
#
# # test end
#
#
#
#
#
#
#
#
#
# model_p = Sequential()
# model_p.add(BatchNormalization())
# model_p.add(LSTM(lstm_size, return_sequences=True))# , input_shape=(timesteps, data_dim)))
# model_p.add(BatchNormalization())
# model_p.add(LSTM(lstm_size))
# # model_p.add(Dense(nb_classes, activation='softmax'))
#
# output_p = model_p(input_p)
#
# auxiliary_output = Dense(nb_classes, activation='softmax', name='aux_output')(output_p)
#
# model_aux = Model(inputs=input_p, outputs=auxiliary_output)
#
# model_aux.compile(optimizer='rmsprop',
#               loss='categorical_crossentropy',
#               metrics=['acc'])
#
# model_aux.fit(train_data_p_lstm, train_label,
#           epochs=30,
#           batch_size=10000,
#           validation_data=[valid_data_p_lstm, valid_label],
#           callbacks=[tb_acc])  # starts training
#
#
#
#
# merged = keras.layers.concatenate([input_f, output_p])
#
# model_final = Sequential()
#
# model_final.add(Dense(256, activation='linear', kernel_initializer='glorot_normal'))
# model_final.add(LeakyReLU())
# # model_final.add(NALU(10))
# model_final.add(BatchNormalization())
# model_final.add(Dense(128, activation='linear', kernel_initializer='glorot_normal'))
# model_final.add(LeakyReLU())
# # model_final.add(NALU(10))
# model_final.add(BatchNormalization())
# model_final.add(Dense(64, activation='linear', kernel_initializer='glorot_normal'))
# model_final.add(LeakyReLU())
# # model_final.add(NALU(10))
# model_final.add(BatchNormalization())
# model_final.add(Dense(32, activation='linear', kernel_initializer='glorot_normal'))
# model_final.add(LeakyReLU())
# # model_final.add(NALU(10))
#
# main_output = Dense(3, activation='softmax', name='output')(model_final(merged))
#
# model = Model(inputs=[input_p, input_f], outputs=[auxiliary_output, main_output])
#
#
# model.compile(optimizer='adam',
#               loss={'aux_output':'categorical_crossentropy', 'output':'categorical_crossentropy'},
#               loss_weights={'aux_output': 0.2, 'output': 1.},
#               metrics=['acc'])
#
# model.fit({'input_p': train_data_p_lstm,'input_f': train_data_f}, {'aux_output': train_label, 'output': train_label},
#           epochs=30,
#           batch_size=10000,
#           validation_data=[{'input_p': valid_data_p_lstm, 'input_f': valid_data_f},
#                            {'aux_output': valid_label, 'output': valid_label}],
#           callbacks=[tb_acc, early_stopping])  # starts training

>>>>>>> origin
